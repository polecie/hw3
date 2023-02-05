import datetime
from pathlib import Path

from openpyxl import Workbook, styles


def add_menus_to_sheet(menus) -> str:
    """"""
    wb = Workbook()
    ws = wb.active

    def _format_dimensions(column: str, width: int):
        ws.column_dimensions[column].width = width

    dimensions = list(zip(["A", "B", "C", "D", "E", "F"], [10, 20, 30, 30, 50, 20]))
    [_format_dimensions(column=dimensions[i][0], width=dimensions[i][1]) for i in range(len(dimensions))]

    for count_menu, menu in enumerate(menus, 1):
        ws.append([count_menu, menu["title"], menu["description"], None, None, None])
        for count_submenu, submenu in enumerate(menu.get("submenus", []), 1):
            ws.append([None, count_submenu, submenu["title"], submenu["description"], None, None])
            for count_dish, dish in enumerate(submenu.get("dishes", []), 1):
                ws.append([None, None, count_dish, dish["title"], dish["description"], float(dish["price"])])

    border = styles.Border(
        left=styles.Side(style="thin"),
        right=styles.Side(style="thin"),
        top=styles.Side(style="thin"),
        bottom=styles.Side(style="thin"),
    )
    font = styles.Font(size=12)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.font = font

    def _format_background(row, column, color: str = "00808080") -> None:
        color = styles.colors.Color(rgb=color)
        background = styles.fills.PatternFill(patternType="solid", fgColor=color)
        cell = ws.cell(row=row, column=column)
        cell.fill = background

    row = 1
    for id, item in enumerate(menus, 1):
        for column in range(1, 7):
            _format_background(row, column)
        row += 1
        for id_sub, sub_item in enumerate(item["submenus"], 1):
            for column in range(2, 7):
                _format_background(row, column, color="00C0C0C0")
            row += 1
            for id_dishes, descript in enumerate(sub_item["dishes"], 1):
                row += 1

    date = datetime.datetime.utcnow().strftime("%Y-%m-%d-%H-%M-%S")
    filename = date + ".xlsx"
    path = "/var/lib/data/menus"
    output_path = Path(path)
    output_path.mkdir(parents=True, exist_ok=True)
    wb.save(f"{path}/{filename}")
    wb.close()
    return f"{path}/{filename}"
